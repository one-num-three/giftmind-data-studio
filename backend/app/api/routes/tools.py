"""Small operational tools used by the first GiftMind collection release."""

from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from pathlib import Path
from typing import Annotated, Literal
from uuid import UUID

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import Field
from sqlalchemy import delete, select

from sqlalchemy.ext.asyncio import AsyncSession
from backend.app.api.deps import SessionContext, get_db_session, require_session
from backend.app.core.config import get_settings
from backend.app.models.assets import GiftImage
from backend.app.models.custom_fields import CustomFieldDefinition, GiftCustomFieldValue
from backend.app.models.gift import Gift
from backend.app.models.operations import BackupRecord
from backend.app.schemas.common import APIModel
from backend.app.schemas.gift import GiftCreate
from backend.app.services.gift_type_inference import GiftTypeDecision, infer_gift_type
from backend.app.services.gifts import GiftNotFoundError, _read_gift, create_gift, get_gift

router = APIRouter(prefix="/api", tags=["tools"])
DatabaseSession = Annotated[AsyncSession, Depends(get_db_session)]
ProtectedSession = Annotated[SessionContext, Depends(require_session)]
DEEPSEEK_MODEL = "deepseek-v4-flash"
DEEPSEEK_TIMEOUT_SECONDS = 45


class CustomFieldInput(APIModel):
    machine_key: Annotated[str, Field(pattern=r"^[a-z][a-z0-9_]*$", max_length=128)]
    display_name: Annotated[str, Field(min_length=1, max_length=256)]
    description: str | None = None
    scope: str = "both"
    value_type: str = "text"
    cardinality: str = "single"
    required_mode: str = "never"
    help_text: str | None = None
    ai_policy: str = "suggest"


class AIInput(APIModel):
    canonical_name: Annotated[str, Field(min_length=1, max_length=256)]
    gift_type_code: str = "product"
    current_values: dict[str, object] = Field(default_factory=dict)


class DeepSeekKeyInput(APIModel):
    api_key: Annotated[str, Field(min_length=10, max_length=256, pattern=r"^\S+$")]


class TaobaoActionInput(APIModel):
    action: Literal["click", "type", "press", "drag", "reload"]
    x: float | None = Field(default=None, ge=0, le=3000)
    y: float | None = Field(default=None, ge=0, le=3000)
    end_x: float | None = Field(default=None, ge=0, le=3000)
    end_y: float | None = Field(default=None, ge=0, le=3000)
    text: str | None = Field(default=None, max_length=2000)
    key: str | None = Field(default=None, max_length=64)


class ExtractInput(APIModel):
    url: Annotated[str, Field(min_length=5, max_length=2048)]


@router.get("/settings/deepseek")
async def deepseek_status(request: Request, _auth: ProtectedSession) -> dict[str, object]:
    settings = request.app.state.settings
    return {"configured": bool(settings.deepseek_api_key), "model": DEEPSEEK_MODEL}


@router.get("/status")
async def server_status(request: Request, _auth: ProtectedSession) -> dict[str, object]:
    settings = request.app.state.settings
    taobao = request.app.state.taobao_login.health()
    return {
        "backend": {"status": "ok", "schemaVersion": settings.schema_version},
        "deepseek": {"configured": bool(settings.deepseek_api_key), "model": DEEPSEEK_MODEL},
        "taobao": {
            "enabled": settings.playwright_enabled,
            **taobao,
        },
    }


@router.put("/settings/deepseek")
async def save_deepseek_key(payload: DeepSeekKeyInput, request: Request, _auth: ProtectedSession) -> dict[str, object]:
    env_path = Path(".env")
    lines = env_path.read_text(encoding="utf-8").splitlines() if env_path.exists() else []
    replacement = f"DEEPSEEK_API_KEY={payload.api_key}"
    output: list[str] = []
    replaced = False
    for line in lines:
        if line.startswith("DEEPSEEK_API_KEY="):
            output.append(replacement)
            replaced = True
        else:
            output.append(line)
    if not replaced:
        output.append(replacement)
    env_path.write_text("\n".join(output) + "\n", encoding="utf-8")
    request.app.state.settings.deepseek_api_key = payload.api_key
    get_settings.cache_clear()
    return {"configured": True, "model": DEEPSEEK_MODEL}


@router.post("/extract")
async def extract_page(payload: ExtractInput, request: Request, _auth: ProtectedSession) -> dict[str, object]:
    """抓取公开页面文字内容。淘宝/天猫链接使用已保存的登录状态。"""
    from backend.app.services.source_extraction import extract_public_page as _extract

    get_settings.cache_clear()
    settings = get_settings()
    state_path = settings.taobao_state_path if settings.taobao_state_path.is_file() else None
    async with httpx.AsyncClient(timeout=20) as client:
        return await _extract(
            payload.url,
            client,
            playwright_enabled=settings.playwright_enabled,
            playwright_timeout_ms=settings.playwright_timeout_ms,
            taobao_state_path=state_path,
        )


@router.post("/taobao/login")
async def start_taobao_login(request: Request, _auth: ProtectedSession) -> dict[str, object]:
    try:
        return await request.app.state.taobao_login.start()
    except RuntimeError as error:
        raise HTTPException(status_code=503, detail=str(error)) from error
    except Exception as error:
        raise HTTPException(status_code=502, detail=f"淘宝登录浏览器启动失败：{type(error).__name__}") from error


@router.get("/taobao/login/{session_id}/status")
async def taobao_login_status(session_id: UUID, request: Request, _auth: ProtectedSession) -> dict[str, object]:
    try:
        return await request.app.state.taobao_login.status(session_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="淘宝登录会话已过期") from error


@router.get("/taobao/login/{session_id}/screenshot")
async def taobao_login_screenshot(session_id: UUID, request: Request, _auth: ProtectedSession) -> Response:
    try:
        image = await request.app.state.taobao_login.screenshot(session_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="淘宝登录会话已过期") from error
    return Response(content=image, media_type="image/png", headers={"Cache-Control": "no-store"})


@router.post("/taobao/login/{session_id}/action")
async def taobao_login_action(
    session_id: UUID,
    payload: TaobaoActionInput,
    request: Request,
    _auth: ProtectedSession,
) -> dict[str, object]:
    try:
        return await request.app.state.taobao_login.action(
            session_id,
            payload.action,
            x=payload.x,
            y=payload.y,
            end_x=payload.end_x,
            end_y=payload.end_y,
            text=payload.text,
            key=payload.key,
        )
    except KeyError as error:
        raise HTTPException(status_code=404, detail="淘宝登录会话已过期") from error
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error


@router.post("/taobao/login/{session_id}/complete")
async def complete_taobao_login(session_id: UUID, request: Request, _auth: ProtectedSession) -> dict[str, object]:
    try:
        return await request.app.state.taobao_login.save(session_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="淘宝登录会话已过期") from error
    except ValueError as error:
        raise HTTPException(status_code=409, detail=str(error)) from error


@router.delete("/taobao/login")
async def clear_taobao_login(request: Request, _auth: ProtectedSession) -> dict[str, bool]:
    await request.app.state.taobao_login.clear()
    return {"cleared": True}


@router.get("/custom-fields")
async def list_custom_fields(session: DatabaseSession, _auth: ProtectedSession) -> list[dict]:
    rows = (await session.execute(select(CustomFieldDefinition).order_by(CustomFieldDefinition.created_at))).scalars().all()
    return [{"id": row.id, "machineKey": row.machine_key, "displayName": row.display_name, "description": row.description, "scope": row.scope, "valueType": row.value_type, "cardinality": row.cardinality, "requiredMode": row.required_mode, "helpText": row.help_text, "aiPolicy": row.ai_policy, "state": row.state} for row in rows]


@router.post("/custom-fields", status_code=status.HTTP_201_CREATED)
async def create_custom_field(payload: CustomFieldInput, session: DatabaseSession, _auth: ProtectedSession) -> dict:
    row = CustomFieldDefinition(**payload.model_dump(), introduced_version=1)
    session.add(row)
    await session.commit()
    return {"id": row.id, "machineKey": row.machine_key, "displayName": row.display_name, "state": row.state}


@router.put("/gifts/{gift_id}/custom-fields")
async def save_custom_values(gift_id: UUID, values: dict[str, object], session: DatabaseSession, _auth: ProtectedSession) -> dict:
    try:
        await get_gift(session, gift_id)
    except GiftNotFoundError as exc:
        raise HTTPException(status_code=404, detail="Gift not found") from exc
    definitions = {row.machine_key: row for row in (await session.execute(select(CustomFieldDefinition).where(CustomFieldDefinition.state == "active"))).scalars().all()}
    await session.execute(delete(GiftCustomFieldValue).where(GiftCustomFieldValue.gift_id == str(gift_id)))
    for key, value in values.items():
        definition = definitions.get(key)
        if definition:
            session.add(GiftCustomFieldValue(gift_id=str(gift_id), field_definition_id=definition.id, value_json=value))
    await session.commit()
    return {"saved": len(values)}


@router.get("/gifts/{gift_id}/custom-fields")
async def get_custom_values(gift_id: UUID, session: DatabaseSession, _auth: ProtectedSession) -> dict[str, object]:
    rows = (await session.execute(select(GiftCustomFieldValue, CustomFieldDefinition).join(CustomFieldDefinition, CustomFieldDefinition.id == GiftCustomFieldValue.field_definition_id).where(GiftCustomFieldValue.gift_id == str(gift_id)))).all()
    return {definition.machine_key: value.value_json for value, definition in rows}


@router.post("/ai/suggest")
async def suggest_with_ai(payload: AIInput, session: DatabaseSession, _auth: ProtectedSession) -> dict:
    del session
    get_settings.cache_clear()
    settings = get_settings()
    key = settings.deepseek_api_key
    selected_type = payload.gift_type_code if payload.gift_type_code in {"product", "activity"} else "product"
    type_evidence = " ".join(
        part
        for part in (
            payload.canonical_name,
            str(payload.current_values.get("shortDescription") or ""),
            str(payload.current_values.get("recipientTypes") or ""),
            str(payload.current_values.get("relationshipStages") or ""),
        )
        if part
    )
    type_decision = infer_gift_type(type_evidence, selected_type)
    expected_type = type_decision.code
    suggestion = _fallback_suggestion(payload.canonical_name, expected_type, type_decision.reason)
    if key:
        prompt = _suggestion_prompt(expected_type, type_decision)
        try:
            async with httpx.AsyncClient(timeout=DEEPSEEK_TIMEOUT_SECONDS) as client:
                response = await client.post("https://api.deepseek.com/chat/completions", headers={"Authorization": f"Bearer {key}"}, json={"model": DEEPSEEK_MODEL, "thinking": {"type": "disabled"}, "temperature": 0.1, "messages": [{"role": "system", "content": prompt}, {"role": "user", "content": json.dumps({"selectedType": selected_type, "expectedType": expected_type, "name": payload.canonical_name, "currentValues": payload.current_values}, ensure_ascii=False)}]})
                response.raise_for_status()
                content = response.json()["choices"][0]["message"]["content"]
                suggestion = _normalize_ai_suggestion(_parse_json_object(content), suggestion, expected_type, type_decision.reason)
        except (httpx.HTTPError, KeyError, TypeError, ValueError, json.JSONDecodeError):
            pass
    return suggestion


def _fallback_suggestion(canonical_name: str, selected_type: str, type_reason: str | None = None) -> dict[str, object]:
    if selected_type == "product":
        purchase_or_booking_tip = "确认配送、兑换和到货时间，尽量保留收到礼物时的惊喜感。"
        ritual_tip = "可以先从对方近期的需要或兴趣自然铺垫，交付时再说明这是专门为他准备的。"
        pairing_ideas = "可配一张手写贺卡，写下具体的喜欢与祝福；也可以在快递备注中留一句简短寄语。"
    else:
        purchase_or_booking_tip = "先确认活动档期、地点、预约规则和取消政策，再把选择权留给对方。"
        ritual_tip = "先表达想和对方一起创造一段回忆，再用轻松语气发出邀请，明确时间可以一起调整。"
        pairing_ideas = "可写一张活动邀请函或承诺券，例如“这次想把时间留给我们，你方便时我们一起去”。"
    return {
        "recommendedGiftTypeCode": selected_type,
        "typeReason": type_reason or "按当前选择的礼物类型生成建议，请人工确认。",
        "subcategoryCode": "other",
        "shortDescription": f"与{canonical_name}相关的礼物或体验。",
        "whyTemplate": f"可以考虑把{canonical_name}送给合适的对象，具体适配关系和价格请人工确认。",
        "priceMin": None,
        "priceMax": None,
        "isFree": False,
        "recipientTypes": [],
        "relationshipStages": [],
        "ageRanges": [],
        "traits": [],
        "interests": [],
        "occasions": [],
        "desiredFeelings": [],
        "memoryHooks": [],
        "tags": [],
        "customTags": [],
        "bestScenarios": None,
        "unsuitableScenarios": None,
        "purchaseOrBookingTip": purchase_or_booking_tip,
        "ritualTip": ritual_tip,
        "pairingIdeas": pairing_ideas,
        "confidence": 0.35,
        "source": "rule",
        "productDetails": {},
        "activityDetails": {},
    }


def _suggestion_prompt(selected_type: str, type_decision: GiftTypeDecision | None = None) -> str:
    type_specific = (
        "productDetails: { productForm, genericProductName, materials[], colors[], sizes[], personalizationMethods[], shippingRequired, digitalDeliveryMethod, isConsumable, shelfLifeDays, storageRequirements, warrantyExpectation }"
        if selected_type == "product"
        else "activityDetails: { activityMode, activityCategory, serviceRegions[], durationMinutesMin, durationMinutesMax, participantsMin, participantsMax, pricingUnit, bookingRequired, validityDays, includedItems[], excludedItems[], ageRestrictions, indoorOutdoor, weatherDependency, cancellationExpectation, refundExpectation }"
    )
    type_reason = type_decision.reason if type_decision else "按当前选择的类型处理。"
    return f"""You are a professional AI gift advisor helping a Chinese gift-data collection team. Return one valid JSON object only, without Markdown fences or extra commentary. The hard expected type is {selected_type}. The type decision is: {type_reason}.

The boundary between Goods and Activity has two hard requirements: the gift giver participates together with the recipient, and their relationship is intimate enough for that shared time to make sense. Goods means the giver does not need to show up and, after delivery, the recipient owns or uses it alone. Goods includes physical products, single-person experiences or services such as a solo spa voucher, individual diving lesson, personal gym card, or electronic redemption code. For Goods, advise on creating an unboxing or receiving surprise and produce a custom-card or delivery-message direction. Activity means the giver must participate with the recipient, in a two-person or group experience such as shared camping, pottery, a concert, a meal, or an escape room, and the relationship should be a close one such as partners, close friends, or family. For Activity, advise on a friendly invitation, schedule coordination, avoiding social pressure, and an invitation letter or promise-coupon direction.

Activity has two hard requirements: the giver and recipient participate together, and their relationship is intimate enough for this shared time to make sense. Look for shared participation such as 一起、共同、双人、多人、陪你、相约 or 邀请, plus an intimate relationship such as 好朋友、好友、闺蜜、伴侣、情侣、恋人、家人、父母或子女. Do not treat 同事、客户、普通合作关系 or an unspecified “朋友” as sufficient proof of intimacy. If either participation or intimacy is missing, do not invent it; keep the selected type and ask the collector to confirm. If the recipient is explicitly alone, or the giver is only sending a voucher or item, use Goods.

Use the guidance fields according to the type: for Goods, purchaseOrBookingTip should cover delivery, timing, unboxing, or redemption; ritualTip should explain a natural surprise setup; pairingIdeas should include 2-3 short card or delivery-message directions. For Activity, purchaseOrBookingTip should cover booking and coordinating dates; ritualTip should explain how to invite without pressure; pairingIdeas should include 2-3 invitation-letter, promise-coupon, or opening-script directions. Do not fill activity-only guidance for Goods or goods-only guidance for Activity.

Type is a hard constraint, not a soft suggestion: recommendedGiftTypeCode MUST be exactly {selected_type}. Only use activity when the deterministic type decision confirms both shared participation and an intimate relationship; for example, “女朋友一起露营”“和好朋友一起观星”“陪父母一起旅行”. A phrase such as “露营”“双人观星” or “和朋友一起” without a close relationship is not enough; use product for automatic inference and ask for confirmation. If the experience is for the recipient alone, use product. If the input describes an item people buy, own, consume, ship, or use (for example 礼盒、冰箱贴、书签、镜头、帐篷、底料、茶叶), use product and fill productDetails. Never return the opposite type just because the UI's previous selection was different. Activity and product details are mutually exclusive: do not put booking, duration, participants, or service regions into productDetails; do not put materials, shipping, sizes, or generic product names into activityDetails.

Do not invent a merchant, URL, exact address, or unverifiable factual claim. Price is an estimated CNY range and may be null when uncertain. Use short Chinese values and only suggest facts that can reasonably be inferred from the name. whyTemplate must contain 4-6 distinct bullet points separated by newlines, with every line starting with '- '; cover the gift's features, likely recipient, suitable occasion, and emotional or practical value without repeating yourself. Never use placeholders such as {{recipient}}, {{occasion}}, or {{relationship}}. Fill every applicable field, and use null or [] when unknown.

Return these keys: recommendedGiftTypeCode (product|activity), typeReason, subcategoryCode, shortDescription, whyTemplate, priceMin, priceMax, isFree, recipientTypes[], relationshipStages[], ageRanges[], traits[], interests[], occasions[], desiredFeelings[], memoryHooks[], tags[], customTags[], bestScenarios, unsuitableScenarios, purchaseOrBookingTip, ritualTip, pairingIdeas, confidence (0 to 1), and {type_specific}."""


def _parse_json_object(content: str) -> dict[str, object]:
    text = content.strip()
    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, flags=re.IGNORECASE | re.DOTALL)
    if fenced:
        text = fenced.group(1).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError("DeepSeek did not return a JSON object")
    parsed = json.loads(text[start : end + 1])
    if not isinstance(parsed, dict):
        raise ValueError("DeepSeek returned a non-object JSON value")
    return parsed


def _string_list(value: object) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item).strip() for item in value if str(item).strip()][:20]


def _nullable_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"\{(?:recipient|receiver|relationship|gift_recipient)\}", "收礼人", text, flags=re.IGNORECASE)
    text = re.sub(r"\{occasion\}", "节日、聚餐或朋友聚会", text, flags=re.IGNORECASE)
    return text or None


def _format_reason_points(text: str) -> str:
    lines = [line.strip() for line in re.split(r"\r?\n+", text) if line.strip()]
    if len(lines) == 1:
        lines = [part.strip() for part in re.split(r"(?<=[。！？；])\s*", lines[0]) if part.strip()]
    lines = [re.sub(r"^(?:[-*•]\s*|\d+[.)、]\s*)", "", line) for line in lines[:6]]
    if 4 <= len(lines) <= 6:
        return "\n".join(f"- {line}" for line in lines)
    return text


def _number(value: object) -> int | float | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number < 0:
        return None
    return int(number) if number.is_integer() else round(number, 2)


def _normalize_product_details(value: object) -> dict[str, object]:
    if not isinstance(value, dict):
        return {}
    result: dict[str, object] = {}
    for key in ("productForm", "genericProductName", "variantNotes", "packageDimensions", "sizeClass", "storageRequirements", "personalizationRequirements", "digitalDeliveryMethod", "shippingNotes", "returnRiskNotes", "warrantyExpectation"):
        if key in value:
            result[key] = value[key] if key == "productForm" else _nullable_text(value[key])
    for key in ("materials", "colors", "sizes", "personalizationMethods", "deviceOrPlatformCompatibility"):
        if key in value:
            result[key] = _string_list(value[key])
    for key in ("weightGrams", "shelfLifeDays"):
        if key in value:
            result[key] = _number(value[key])
    for key in ("isBulky", "isFragile", "isConsumable", "shippingRequired"):
        if key in value and isinstance(value[key], bool):
            result[key] = value[key]
    return result


def _normalize_activity_details(value: object) -> dict[str, object]:
    if not isinstance(value, dict):
        return {}
    result: dict[str, object] = {}
    for key in ("activityMode", "activityCategory", "pricingUnit", "scheduleType", "equipmentRequirements", "ageRestrictions", "heightRestrictions", "healthRestrictions", "accessibilityNotes", "weatherDependency", "indoorOutdoor", "cancellationExpectation", "rescheduleExpectation", "refundExpectation"):
        if key in value:
            result[key] = value[key] if key == "activityMode" else _nullable_text(value[key])
    for key in ("serviceRegions", "includedItems", "excludedItems"):
        if key in value:
            result[key] = _string_list(value[key])
    for key in ("durationMinutesMin", "durationMinutesMax", "participantsMin", "participantsMax", "bookingLeadDaysMin", "bookingLeadDaysMax", "validityDays"):
        if key in value:
            result[key] = _number(value[key])
    if "bookingRequired" in value and isinstance(value["bookingRequired"], bool):
        result["bookingRequired"] = value["bookingRequired"]
    return result


def _normalize_ai_suggestion(
    raw: dict[str, object],
    fallback: dict[str, object],
    selected_type: str,
    type_reason: str | None = None,
) -> dict[str, object]:
    result = dict(fallback)
    for key in ("typeReason", "subcategoryCode", "shortDescription", "whyTemplate", "bestScenarios", "unsuitableScenarios", "purchaseOrBookingTip", "ritualTip", "pairingIdeas"):
        if key in raw:
            value = _nullable_text(raw[key])
            if key == "whyTemplate" and value:
                value = _format_reason_points(value)
            result[key] = value or result[key]
    # The model cannot override the deterministic type decision. This also
    # protects the form when DeepSeek follows a stale/default UI selection.
    result["recommendedGiftTypeCode"] = selected_type
    result["typeReason"] = type_reason or result.get("typeReason") or "按当前选择的礼物类型生成建议，请人工确认。"
    for key in ("recipientTypes", "relationshipStages", "ageRanges", "traits", "interests", "occasions", "desiredFeelings", "memoryHooks", "tags", "customTags"):
        if key in raw:
            result[key] = _string_list(raw[key])
    for key in ("priceMin", "priceMax"):
        if key in raw:
            result[key] = _number(raw[key])
    if isinstance(raw.get("isFree"), bool):
        result["isFree"] = raw["isFree"]
    if result["isFree"]:
        result["priceMin"] = None
        result["priceMax"] = None
    elif result["priceMin"] is not None and result["priceMax"] is not None and result["priceMin"] > result["priceMax"]:
        result["priceMin"], result["priceMax"] = result["priceMax"], result["priceMin"]
    confidence = _number(raw.get("confidence"))
    if confidence is not None:
        result["confidence"] = min(1, confidence)
    if selected_type == "product":
        result["productDetails"] = _normalize_product_details(raw.get("productDetails"))
        result["activityDetails"] = {}
    else:
        result["productDetails"] = {}
        result["activityDetails"] = _normalize_activity_details(raw.get("activityDetails"))
    result["source"] = "deepseek"
    return result


@router.post("/gifts/{gift_id}/images", status_code=201)
async def upload_gift_image(gift_id: UUID, session: DatabaseSession, _auth: ProtectedSession, file: UploadFile = File(...)) -> dict:
    settings = get_settings()
    content = await file.read()
    if len(content) > 8 * 1024 * 1024 or (file.content_type or "") not in {"image/jpeg", "image/png", "image/webp"}:
        raise HTTPException(status_code=422, detail="仅支持 8MB 内的 JPG、PNG、WebP 图片")
    upload_dir = Path(settings.upload_dir); upload_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(content).hexdigest(); suffix = Path(file.filename or "image.jpg").suffix.lower() or ".jpg"; stored = f"{digest}{suffix}"; (upload_dir / stored).write_bytes(content)
    row = GiftImage(gift_id=str(gift_id), original_filename=file.filename or stored, stored_filename=stored, content_type=file.content_type or "image/jpeg", sha256=digest, file_size_bytes=len(content))
    session.add(row); await session.commit()
    return {"id": row.id, "filename": row.original_filename, "url": f"/uploads/{stored}", "isCover": row.is_cover}


@router.get("/gifts/{gift_id}/images")
async def list_gift_images(gift_id: UUID, session: DatabaseSession, _auth: ProtectedSession) -> list[dict]:
    rows = (await session.execute(select(GiftImage).where(GiftImage.gift_id == str(gift_id)).order_by(GiftImage.display_order, GiftImage.created_at))).scalars().all()
    return [{"id": row.id, "filename": row.original_filename, "url": f"/uploads/{row.stored_filename}", "isCover": row.is_cover} for row in rows]


@router.delete("/images/{image_id}", status_code=200)
async def delete_gift_image(image_id: UUID, session: DatabaseSession, _auth: ProtectedSession) -> dict[str, bool]:
    row = (await session.execute(select(GiftImage).where(GiftImage.id == str(image_id)))).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Image not found")
    path = get_settings().upload_dir / row.stored_filename
    if path.exists(): path.unlink()
    await session.delete(row); await session.commit(); return {"deleted": True}


@router.get("/export/xlsx")
async def export_xlsx(session: DatabaseSession, _auth: ProtectedSession) -> StreamingResponse:
    gifts = (await session.execute(select(Gift).where(Gift.deleted_at.is_(None)).order_by(Gift.canonical_name))).scalars().all()
    book = Workbook(); sheet = book.active; sheet.title = "gifts"; sheet.append(["id", "name", "type", "status", "description", "price_min", "price_max", "tags"])
    for gift in gifts: sheet.append([gift.id, gift.canonical_name, gift.gift_type_code, gift.status, gift.short_description or "", gift.price_min, gift.price_max, ", ".join(gift.tags or [])])
    output = io.BytesIO(); book.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=giftmind-gifts.xlsx"})


@router.post("/import/xlsx")
async def import_xlsx(session: DatabaseSession, _auth: ProtectedSession, file: UploadFile = File(...)) -> dict:
    book = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True); sheet = book.active; rows = list(sheet.iter_rows(values_only=True)); imported = 0; rejected = []
    for number, row in enumerate(rows[1:], start=2):
        if not row or not row[1]: continue
        try:
            payload = {"canonicalName": str(row[1]).strip(), "giftTypeCode": row[2] or "product", "status": row[3] or "draft", "shortDescription": row[4], "priceMin": row[5], "priceMax": row[6], "tags": [item.strip() for item in str(row[7] or "").split(",") if item.strip()]}
            await create_gift(session, GiftCreate.model_validate(payload)); imported += 1
        except Exception as exc:
            rejected.append({"row": number, "error": str(exc)})
    return {"total": max(0, len(rows) - 1), "imported": imported, "rejected": rejected}


@router.get("/backup")
async def download_backup(session: DatabaseSession, _auth: ProtectedSession) -> StreamingResponse:
    gifts = (await session.execute(select(Gift).where(Gift.deleted_at.is_(None)))).scalars().all()
    payload = {"schemaVersion": 1, "gifts": [{"name": row.canonical_name, "type": row.gift_type_code, "status": row.status, "description": row.short_description, "tags": row.tags} for row in gifts]}
    data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"); output = io.BytesIO();
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive: archive.writestr("giftmind-backup.json", data)
    output.seek(0); return StreamingResponse(output, media_type="application/zip", headers={"Content-Disposition": "attachment; filename=giftmind-backup.zip"})


@router.post("/restore")
async def restore_backup(session: DatabaseSession, _auth: ProtectedSession, file: UploadFile = File(...)) -> dict:
    raw = await file.read()
    with zipfile.ZipFile(io.BytesIO(raw)) as archive: payload = json.loads(archive.read("giftmind-backup.json"))
    imported = 0
    for item in payload.get("gifts", []):
        try:
            await create_gift(session, GiftCreate.model_validate({"canonicalName": item["name"], "giftTypeCode": item.get("type", "product"), "status": "draft", "shortDescription": item.get("description"), "tags": item.get("tags", [])})); imported += 1
        except Exception: continue
    return {"restored": imported}
